#!/usr/bin/env python3
"""
Gemma Realtime - File Analyzer
Multi-format file analysis (code, images, PDFs, docs, DB, audio, video)
"""

import io
import os
import json
import asyncio
import subprocess
from typing import Optional, Dict, Any, List
from pathlib import Path
import torch
from PIL import Image

# Code file extensions
CODE_EXTENSIONS = {
    '.py', '.js', '.ts', '.jsx', '.tsx', '.java', '.c', '.cpp', '.h', '.hpp',
    '.cs', '.go', '.rs', '.rb', '.php', '.swift', '.kt', '.scala', '.lua',
    '.sh', '.bash', '.zsh', '.ps1', '.bat', '.cmd', '.sql', '.r', '.m',
    '.html', '.css', '.scss', '.sass', '.less', '.vue', '.svelte',
    '.json', '.yaml', '.yml', '.toml', '.xml', '.ini', '.env',
    '.md', '.txt', '.rst', '.tex',
}

# Image extensions
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.svg'}

# Document extensions
DOC_EXTENSIONS = {'.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt'}

# Spreadsheet extensions  
XLSX_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.ods'}

# Database extensions
DB_EXTENSIONS = {'.db', '.sqlite', '.sqlite3'}

# Audio extensions
AUDIO_EXTENSIONS = {'.mp3', '.wav', '.ogg', '.flac', '.m4a', '.aac', '.opus'}

# Video extensions
VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mkv', '.mov', '.webm', '.flv'}

class FileAnalyzer:
    """Multi-format file analyzer"""
    
    def __init__(self):
        self.vision_model = None
        self.redis_client = None
    
    async def load(self):
        """Initialize components"""
        print("[FileAnalyzer] Loading...")
        
        # Try to load vision model
        try:
            # Will use existing Ollama with vision capabilities
            self.vision_available = True
        except:
            self.vision_available = False
        
        print(f"[FileAnalyzer] Vision available: {self.vision_available}")
    
    async def analyze(self, file_data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze a file and return description"""
        
        ext = Path(filename).suffix.lower()
        
        # Route to appropriate analyzer
        if ext in CODE_EXTENSIONS:
            return await self._analyze_code(file_data, filename)
        elif ext in IMAGE_EXTENSIONS:
            return await self._analyze_image(file_data, filename)
        elif ext in DOC_EXTENSIONS:
            return await self._analyze_pdf(file_data, filename)
        elif ext in XLSX_EXTENSIONS:
            return await self._analyze_spreadsheet(file_data, filename)
        elif ext in DB_EXTENSIONS:
            return await self._analyze_database(file_data, filename)
        elif ext in AUDIO_EXTENSIONS:
            return await self._analyze_audio(filename)
        elif ext in VIDEO_EXTENSIONS:
            return await self._analyze_video(filename)
        else:
            return await self._analyze_generic(file_data, filename)
    
    async def _analyze_code(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze code files"""
        try:
            text = data.decode('utf-8', errors='ignore')
            lines = text.split('\n')
            
            # Basic analysis
            analysis = {
                "type": "code",
                "filename": filename,
                "lines": len(lines),
                "size_bytes": len(data),
                "analysis": {
                    "has_imports": "import " in text or "require" in text,
                    "has_functions": "def " in text or "function " in text,
                    "has_classes": "class " in text,
                    "has_docstrings": '"""' in text or "'''" in text,
                }
            }
            
            # Try to extract more info
            if "def " in text:
                funcs = [l.strip() for l in lines if "def " in l and ":" in l]
                analysis["analysis"]["functions"] = funcs[:10]  # First 10
            
            if "class " in text:
                classes = [l.strip() for l in lines if "class " in l and ":" in l]
                analysis["analysis"]["classes"] = classes[:10]
            
            # Summary for LLM
            analysis["summary"] = f"""Code file: {filename}
- {len(lines)} lines
- {len(analysis['analysis'].get('functions', []))} functions
- {len(analysis['analysis'].get('classes', []))} clases
First 500 chars: {text[:500]}"""
            
            return analysis
            
        except Exception as e:
            return {"type": "code", "error": str(e)}
    
    async def _analyze_image(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze images with vision"""
        result = {
            "type": "image",
            "filename": filename,
            "size_bytes": len(data),
        }
        
        # Get basic image info
        try:
            img = Image.open(io.BytesIO(data))
            result["image_info"] = {
                "format": img.format,
                "size": img.size,
                "mode": img.mode,
            }
            result["summary"] = f"Image: {filename}, {img.size[0]}x{img.size[1]}, {img.format}"
        except Exception as e:
            result["summary"] = f"Image file: {filename} ({len(data)} bytes)"
        
        # Try vision if available
        if self.vision_available:
            try:
                # Use Ollama with vision
                import base64
                b64 = base64.b64encode(data).decode()
                
                # This would call Ollama with vision
                result["needs_vision"] = True
                result["summary"] += "\n[Needs vision model for detailed analysis]"
            except:
                pass
        
        return result
    
    async def _analyze_pdf(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze PDF files"""
        result = {
            "type": "document",
            "filename": filename,
            "size_bytes": len(data),
        }
        
        try:
            import pdfplumber
            
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                result["pdf_info"] = {
                    "pages": len(pdf.pages),
                }
                
                # Extract text from first few pages
                text_parts = []
                for i, page in enumerate(pdf.pages[:3]):
                    text = page.extract_text()
                    if text:
                        text_parts.append(text[:500])
                
                full_text = "\n".join(text_parts)
                result["summary"] = f"""PDF: {filename}
- {len(pdf.pages)} pages
Content preview:
{full_text}"""
                
        except Exception as e:
            result["summary"] = f"PDF file: {filename} ({len(data)} bytes)\nError: {e}"
        
        return result
    
    async def _analyze_spreadsheet(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze spreadsheets"""
        result = {
            "type": "spreadsheet",
            "filename": filename,
            "size_bytes": len(data),
        }
        
        try:
            import openpyxl
            if filename.endswith('.csv'):
                result["summary"] = f"CSV file: {filename} ({len(data)} bytes)"
            else:
                wb = openpyxl.load_workbook(io.BytesIO(data))
                result["excel_info"] = {
                    "sheets": wb.sheetnames,
                    "sheet_count": len(wb.sheetnames)
                }
                
                # Get first sheet info
                ws = wb.active
                result["summary"] = f"""Excel: {filename}
- {len(wb.sheetnames)} sheets
- Sheet '{ws.title}': {ws.max_row} rows x {ws.max_column} cols
Sample data (first 3 rows):"""
                
        except Exception as e:
            result["summary"] = f"Spreadsheet: {filename} ({len(data)} bytes)\nError: {e}"
        
        return result
    
    async def _analyze_database(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Analyze database files"""
        result = {
            "type": "database",
            "filename": filename,
            "size_bytes": len(data),
        }
        
        # For SQLite, just provide basic info
        # Full analysis would require SQL queries
        result["summary"] = f"""Database file: {filename}
- Size: {len(data)} bytes
- Type: SQLite
- Note: Contains structured data. Query for schema and sample data."""
        
        return result
    
    async def _analyze_audio(self, filename: str) -> Dict[str, Any]:
        """Analyze audio files"""
        result = {
            "type": "audio",
            "filename": filename,
            "summary": f"Audio file: {filename}\n- Needs audio transcription service for content analysis."
        }
        
        return result
    
    async def _analyze_video(self, filename: str) -> Dict[str, Any]:
        """Analyze video files"""
        result = {
            "type": "video", 
            "filename": filename,
            "needs_processing": True,
            "summary": f"""Video file: {filename}
- This is a video file.
- For analysis, I can:
  A) Extract keyframes and describe them
  B) Transcribe audio track
  C) Both
Which would you prefer?""",
            "options": [
                {"id": "keyframes", "label": "Extract keyframes"},
                {"id": "transcribe", "label": "Transcribe audio"},
                {"id": "both", "label": "Both"}
            ]
        }
        
        return result
    
    async def _analyze_generic(self, data: bytes, filename: str) -> Dict[str, Any]:
        """Generic file analysis"""
        return {
            "type": "generic",
            "filename": filename,
            "size_bytes": len(data),
            "summary": f"File: {filename} ({len(data)} bytes)"
        }
    
    async def ask_video_choice(self, filename: str, choice: str) -> Dict[str, Any]:
        """Process user's video analysis choice"""
        result = {
            "type": "video",
            "filename": filename,
            "choice": choice
        }
        
        if choice == "keyframes":
            result["summary"] = f"Extracting keyframes from {filename}..."
            # Would use ffmpeg to extract keyframes
        elif choice == "transcribe":
            result["summary"] = f"Transcribing audio from {filename}..."
            # Would use Whisper on audio track
        elif choice == "both":
            result["summary"] = f"Extracting keyframes and transcribing audio from {filename}..."
        
        return result

# Global instance
file_analyzer = FileAnalyzer()

async def init_analyzer():
    """Initialize analyzer"""
    await file_analyzer.load()
    print("[FileAnalyzer] Ready")

if __name__ == "__main__":
    asyncio.run(init_analyzer())